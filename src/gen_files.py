# -*- coding: utf-8 -*-
"""生成 CSV / XLSX / .ase / PowerPoint 主题色 XML / GIMP .gpl / Origin 兼容 txt。"""
import json, os, struct, csv
from colorlib import hex2rgb, lch, contrast

lib = json.load(open('library.json'))
OUT = 'build'
os.makedirs(OUT, exist_ok=True)
os.makedirs(f'{OUT}/ase', exist_ok=True)
os.makedirs(f'{OUT}/ppt-theme-colors', exist_ok=True)

SLOT = ['1 主', '2 次', '3 三', '4 四', '5 五', '6 六']

# ----------------------------------------------------------------- CSV
rows = []
for e in lib:
    for i, c in enumerate(e['colors']):
        r, g, b = [round(v * 255) for v in hex2rgb(c)]
        L, C, h = lch(c)
        rows.append(dict(
            slug=e['slug'], 中文名=e['name_zh'], 英文名=e['name_en'],
            角色=e['zh'], 色调=e['tone_zh'], 色系=e['family'], 出处=e['source'],
            序号=i + 1, HEX=c, R=r, G=g, B=b,
            RGB=f'{r},{g},{b}', **{'L*': round(L, 1), 'C*': round(C, 1), 'h°': round(h, 1)},
            白底对比度=round(contrast(c, '#FFFFFF'), 2),
            平滑序=e['orders']['smooth'].index(i) + 1,
            区分度序=e['orders']['distinct'].index(i) + 1,
            色相序=e['orders']['hue'].index(i) + 1,
            明度序=e['orders']['light'].index(i) + 1,
            色盲等级=e['cvd_grade'],
            色盲安全=('是' if i in e['safe_set'] else '否'),
            深色变体=e['dark'][i], 浅色变体=e['light'][i],
            纸色bg=e['bg'], 墨色ink=e['ink'],
        ))
with open(f'{OUT}/anime_palettes.csv', 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    w.writeheader()
    w.writerows(rows)
print('csv ok', len(rows), 'rows')

# ----------------------------------------------------------------- XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- sheet 1: 色卡总览
ws = wb.active
ws.title = '色卡总览'
hdr = ['slug', '中文名', '英文名', '色系', '出处', '色盲等级',
       '色1', '色2', '色3', '色4', '色5', '色6',
       'HEX 1（平滑序）', 'HEX 2', 'HEX 3', 'HEX 4', 'HEX 5', 'HEX 6',
       '纸色 bg', '墨色 ink', '色盲安全序号', '最小 ΔE00']
ws.append(hdr)
thin = Side(style='thin', color='D9D9D9')
for j, _ in enumerate(hdr, 1):
    c = ws.cell(1, j)
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='3B3B44')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for e in lib:
    r = ws.max_row + 1
    vals = ([e['slug'], e['name_zh'], e['name_en'], e['family'], e['source'], e['cvd_grade']]
            + [''] * 6 + e['colors'] + [e['bg'], e['ink'],
            ' '.join(str(i + 1) for i in e['safe_set']), e['min_de']])
    for j, v in enumerate(vals, 1):
        cell = ws.cell(r, j, v)
        cell.font = Font(name='Arial', size=9)
        cell.border = Border(thin, thin, thin, thin)
    for k in range(6):                              # 色块
        cell = ws.cell(r, 7 + k)
        cell.fill = PatternFill('solid', fgColor=e['colors'][k][1:])
    for k in range(6):                              # HEX 文本，底色同色
        cell = ws.cell(r, 13 + k)
        cell.fill = PatternFill('solid', fgColor=e['colors'][k][1:])
        cell.font = Font(name='Arial', size=9, bold=True,
                         color='FFFFFF' if contrast(e['colors'][k], '#FFFFFF') > 3.2 else '111111')
        cell.alignment = Alignment(horizontal='center')
    ws.cell(r, 19).fill = PatternFill('solid', fgColor=e['bg'][1:])
    ws.cell(r, 20).fill = PatternFill('solid', fgColor=e['ink'][1:])
    ws.cell(r, 20).font = Font(name='Arial', size=9, color='FFFFFF')
widths = [24, 18, 26, 6, 16, 8] + [5] * 6 + [10] * 6 + [10, 10, 14, 10]
for j, wd in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = wd
ws.freeze_panes = 'C2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(hdr))}{ws.max_row}'

# --- sheet 2: 明细（每色一行）
ws2 = wb.create_sheet('逐色明细')
keys = list(rows[0].keys())
ws2.append(keys)
for j in range(1, len(keys) + 1):
    c = ws2.cell(1, j)
    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='3B3B44')
    c.alignment = Alignment(horizontal='center', wrap_text=True)
hexcol = keys.index('HEX') + 1
for row in rows:
    ws2.append([row[k] for k in keys])
    r = ws2.max_row
    for j in range(1, len(keys) + 1):
        ws2.cell(r, j).font = Font(name='Arial', size=9)
    cell = ws2.cell(r, hexcol)
    cell.fill = PatternFill('solid', fgColor=row['HEX'][1:])
    cell.font = Font(name='Arial', size=9, bold=True,
                     color='FFFFFF' if contrast(row['HEX'], '#FFFFFF') > 3.2 else '111111')
for j, k in enumerate(keys, 1):
    ws2.column_dimensions[get_column_letter(j)].width = max(9, min(20, len(k) * 2.1))
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(keys))}{ws2.max_row}'

# --- sheet 3: 深浅延伸
ws3 = wb.create_sheet('深浅延伸')
ws3.append(['slug', '中文名', '', '浅色 light（+16 L*）', '', '', '', '', '',
            '主色 main', '', '', '', '', '', '深色 dark（-15 L*）'])
for j in range(1, 26):
    ws3.cell(1, j).font = Font(name='Arial', bold=True, size=9)
for e in lib:
    r = ws3.max_row + 1
    ws3.cell(r, 1, e['slug']).font = Font(name='Arial', size=9)
    ws3.cell(r, 2, e['name_zh']).font = Font(name='Arial', size=9)
    for k in range(6):
        for off, key in ((3, 'light'), (10, 'colors'), (17, 'dark')):
            cell = ws3.cell(r, off + k, e[key][k])
            cell.fill = PatternFill('solid', fgColor=e[key][k][1:])
            cell.font = Font(name='Arial', size=8,
                             color='FFFFFF' if contrast(e[key][k], '#FFFFFF') > 3.2 else '111111')
            cell.alignment = Alignment(horizontal='center')
ws3.column_dimensions['A'].width = 24
ws3.column_dimensions['B'].width = 18
for j in range(3, 24):
    ws3.column_dimensions[get_column_letter(j)].width = 9
ws3.freeze_panes = 'C2'

# --- sheet 4: 排序方案
ORD_LABEL = [('smooth', '平滑（默认）'), ('distinct', '区分度优先（图表用）'),
             ('hue', '色相环'), ('light', '明度浅→深')]
ws5 = wb.create_sheet('排序方案')
ws5.append(['slug', '中文名', '排序', '1', '2', '3', '4', '5', '6'])
for j in range(1, 10):
    c = ws5.cell(1, j)
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='3B3B44')
    c.alignment = Alignment(horizontal='center')
for e in lib:
    for k, (key, label) in enumerate(ORD_LABEL):
        r = ws5.max_row + 1
        ws5.cell(r, 1, e['slug'] if k == 0 else '').font = Font(name='Arial', size=9)
        ws5.cell(r, 2, e['name_zh'] if k == 0 else '').font = Font(name='Arial', size=9)
        ws5.cell(r, 3, label).font = Font(name='Arial', size=9,
                                          bold=(k == 0), color='1D1D21' if k == 0 else '6C6C76')
        for pos, idx in enumerate(e['orders'][key]):
            col = e['colors'][idx]
            cell = ws5.cell(r, 4 + pos, col)
            cell.fill = PatternFill('solid', fgColor=col[1:])
            cell.font = Font(name='Arial', size=8, bold=True,
                             color='FFFFFF' if contrast(col, '#FFFFFF') > 3.2 else '111111')
            cell.alignment = Alignment(horizontal='center')
ws5.column_dimensions['A'].width = 24
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 20
for j in range(4, 10):
    ws5.column_dimensions[get_column_letter(j)].width = 11
ws5.freeze_panes = 'D2'

# --- sheet 5: 色标采样（连续色标的 11 级取样，可直接复制进 Origin/Excel 条件格式）
RAMPS = [('seq', '连续 seq（单色相：热图/密度）'),
         ('flow', '强过渡 flow（多色相：散点连续着色）'),
         ('div', '发散 div（以 0 为中心）'),
         ('cyclic', '环形 cyclic（相位/角度/时刻）')]
ws6 = wb.create_sheet('色标采样')
ws6.append(['slug', '中文名', '色标'] + [f'{int(i*10)}%' for i in range(11)]
           + ['明度跨度 L*', '明度单调', '感知均匀度'])
for j in range(1, 18):
    c = ws6.cell(1, j)
    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='3B3B44')
    c.alignment = Alignment(horizontal='center', wrap_text=True)
for e in lib:
    for k, (key, label) in enumerate(RAMPS):
        r = ws6.max_row + 1
        ws6.cell(r, 1, e['slug'] if k == 0 else '').font = Font(name='Arial', size=9)
        ws6.cell(r, 2, e['name_zh'] if k == 0 else '').font = Font(name='Arial', size=9)
        ws6.cell(r, 3, label).font = Font(name='Arial', size=9)
        for t in range(11):
            col = e[key][round(t * 255 / 10)]
            cell = ws6.cell(r, 4 + t, col)
            cell.fill = PatternFill('solid', fgColor=col[1:])
            cell.font = Font(name='Arial', size=7.5, bold=True,
                             color='FFFFFF' if contrast(col, '#FFFFFF') > 3.2 else '111111')
            cell.alignment = Alignment(horizontal='center')
        st = e['ramp_stats'][key]
        for off, v in enumerate((st['L_range'], '是' if st['monotonic'] else '否',
                                 st['uniformity'])):
            ws6.cell(r, 15 + off, v).font = Font(name='Arial', size=9)
ws6.column_dimensions['A'].width = 24
ws6.column_dimensions['B'].width = 18
ws6.column_dimensions['C'].width = 32
for j in range(4, 15):
    ws6.column_dimensions[get_column_letter(j)].width = 9.5
for j in (15, 16, 17):
    ws6.column_dimensions[get_column_letter(j)].width = 12
ws6.freeze_panes = 'D2'

# --- sheet 6: 使用说明
ws4 = wb.create_sheet('使用说明')
notes = [
    ['动漫 / 游戏角色配色库 v1.0', ''],
    ['', ''],
    ['命名规则', '角色名 · 色调标签（如「胡桃 · 绯梅」），slug 为英文小写连字符形式'],
    ['每套结构', '6 个主色 + 深/浅变体 + 纸色 bg + 次级底 bg2 + 辅助灰 muted + 墨色 ink'],
    ['默认排序', '「平滑」：以角色主色为起点，相邻两色的 ΔE00 之和最小，过渡最顺滑。本表各处默认用这个顺序'],
    ['画图表时', '换成「区分度优先」——前几色彼此差异最大，只画 2–3 条线时直接取前 2–3 色。见「排序方案」表'],
    ['另两种排序', '「色相环」红→橙→黄→绿→青→蓝→紫；「明度浅→深」适合有序数据与灰度打印'],
    ['', ''],
    ['四种连续色标', '见「色标采样」表。seq 单色相；flow 多色相且明度严格单调，'
                     '**散点/密度等强过渡场景用它**；div 发散；cyclic 环形（相位、角度、时刻）'],
    ['散点图建议', '分类散点→用「区分度优先」的主色 + 白描边 + 换形状；'
                   '连续变量着色→用 flow，并裁掉最浅的 10%（太浅的小点在白底上看不见）；'
                   '点特别多→缩小点径、降 alpha、去描边，只用色标较深的一半'],
    ['色盲等级 A', '红/绿色盲模拟下 6 色两两 ΔE00 ≥ 13，整套可直接用于多系列图'],
    ['色盲等级 B', '大部分可分，建议 4 个系列以内，或按「色盲安全序号」取子集'],
    ['色盲等级 C', '角色本身就是同色系，多系列时请只用「色盲安全序号」列出的那几个'],
    ['最小 ΔE00', '正常视觉下同套配色内任意两色的最小色差；> 15 表示区分度充裕'],
    ['白底对比度', 'WCAG 对比度，用作线条/柱体时建议 ≥ 2.0，用作正文文字时 ≥ 4.5'],
    ['', ''],
    ['PPT 用法', '把 ppt-theme-colors 里的 .xml 放进「Document Themes\\Theme Colors」文件夹，'
                 '再从「设计 → 变体 → 颜色」里选；或直接从本表复制 HEX。'
                 '主题里的 accent1–6 用「区分度优先」顺序，因为 PowerPoint 图表是按 accent 顺序循环取色的'],
    ['Python 用法', 'import anime_palettes as ap; ap.use("miku")  → 默认平滑序；'
                    'ap.use("miku", order="distinct") → 图表用；ap.cmap("miku","div") → 发散色标'],
    ['AI/PS 用法', '导入 ase/ 目录下的 .ase 色板文件'],
    ['Origin 用法', 'origin-hex/平滑/ 与 origin-hex/区分度/ 两个目录，各套一个 txt，逐行粘进自定义颜色列表'],
]
for row in notes:
    ws4.append(row)
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 110
for r in range(1, ws4.max_row + 1):
    ws4.cell(r, 1).font = Font(name='Arial', size=10, bold=True)
    ws4.cell(r, 2).font = Font(name='Arial', size=10)
    ws4.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')
ws4['A1'].font = Font(name='Arial', size=13, bold=True)

wb.save(f'{OUT}/anime_palettes.xlsx')
print('xlsx ok')


# ----------------------------------------------------------------- .ase
def _ase_str(s):
    data = s.encode('utf-16-be') + b'\x00\x00'
    return struct.pack('>H', len(s) + 1) + data


def ase_color_block(name, hexcolor):
    r, g, b = hex2rgb(hexcolor)
    body = _ase_str(name) + b'RGB ' + struct.pack('>fff', r, g, b) + struct.pack('>H', 2)
    return struct.pack('>HI', 0x0001, len(body)) + body


def ase_group_start(name):
    body = _ase_str(name)
    return struct.pack('>HI', 0xC001, len(body)) + body


def ase_group_end():
    return struct.pack('>HI', 0xC002, 0)


def write_ase(path, groups):
    blocks, n = b'', 0
    for gname, items in groups:
        if gname:
            blocks += ase_group_start(gname); n += 1
        for cname, chex in items:
            blocks += ase_color_block(cname, chex); n += 1
        if gname:
            blocks += ase_group_end(); n += 1
    with open(path, 'wb') as f:
        f.write(b'ASEF' + struct.pack('>HH', 1, 0) + struct.pack('>I', n) + blocks)


groups_all = []
for e in lib:
    items = [(f"{e['en']} {i+1} {c}", c) for i, c in enumerate(e['colors'])]
    items += [(f"{e['en']} bg {e['bg']}", e['bg']), (f"{e['en']} ink {e['ink']}", e['ink'])]
    groups_all.append((f"{e['zh']} {e['tone_zh']} / {e['en']}", items))
    write_ase(f"{OUT}/ase/{e['slug']}.ase", [(None, items)])
write_ase(f'{OUT}/ase/_ALL-36-anime-palettes.ase', groups_all)
print('ase ok', len(groups_all) + 1, 'files')

# ----------------------------------------------------------------- GIMP .gpl
os.makedirs(f'{OUT}/gpl', exist_ok=True)
for e in lib:
    lines = ['GIMP Palette', f"Name: {e['en']} - {e['tone_en']}", 'Columns: 6', '#']
    for i, c in enumerate(e['colors'] + [e['bg'], e['ink']]):
        r, g, b = [round(v * 255) for v in hex2rgb(c)]
        lines.append(f"{r:3d} {g:3d} {b:3d}\t{e['en']} {i+1}")
    open(f"{OUT}/gpl/{e['slug']}.gpl", 'w').write('\n'.join(lines) + '\n')
print('gpl ok')

# ----------------------------------------------------------------- Origin / 粘贴用
for sub, key in (('平滑', 'smooth'), ('区分度', 'distinct')):
    os.makedirs(f'{OUT}/origin-hex/{sub}', exist_ok=True)
    for e in lib:
        cs = [e['colors'][i] for i in e['orders'][key]]
        open(f"{OUT}/origin-hex/{sub}/{e['slug']}.txt", 'w').write('\n'.join(cs) + '\n')
os.makedirs(f'{OUT}/origin-hex/flow色标', exist_ok=True)
for e in lib:
    cs = [e['flow'][round(t * 255 / 15)] for t in range(16)]
    open(f"{OUT}/origin-hex/flow色标/{e['slug']}.txt", 'w').write('\n'.join(cs) + '\n')
print('origin ok')

# ----------------------------------------------------------------- PPT 主题色 XML
TPL = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<a:clrScheme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="{name}">
  <a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>
  <a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>
  <a:dk2><a:srgbClr val="{dk2}"/></a:dk2>
  <a:lt2><a:srgbClr val="{lt2}"/></a:lt2>
  <a:accent1><a:srgbClr val="{a1}"/></a:accent1>
  <a:accent2><a:srgbClr val="{a2}"/></a:accent2>
  <a:accent3><a:srgbClr val="{a3}"/></a:accent3>
  <a:accent4><a:srgbClr val="{a4}"/></a:accent4>
  <a:accent5><a:srgbClr val="{a5}"/></a:accent5>
  <a:accent6><a:srgbClr val="{a6}"/></a:accent6>
  <a:hlink><a:srgbClr val="{hl}"/></a:hlink>
  <a:folHlink><a:srgbClr val="{fh}"/></a:folHlink>
</a:clrScheme>
'''
for e in lib:
    # accent1–6 用「区分度优先」顺序：PowerPoint 图表按 accent 顺序循环取色
    cs = [e['colors'][i][1:] for i in e['orders']['distinct']]
    xml = TPL.format(name=f"{e['zh']} {e['tone_zh']}", dk2=e['ink'][1:], lt2=e['bg'][1:],
                     a1=cs[0], a2=cs[1], a3=cs[2], a4=cs[3], a5=cs[4], a6=cs[5],
                     hl=e['dark'][0][1:], fh=e['muted'][1:])
    open(f"{OUT}/ppt-theme-colors/{e['slug']}.xml", 'w', encoding='utf-8').write(xml)
print('ppt theme xml ok')

json.dump(lib, open(f'{OUT}/anime_palettes.json', 'w'), ensure_ascii=False, indent=1)
print('json ok')
